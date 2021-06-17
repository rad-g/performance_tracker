import win32gui
import time
import datetime
import threading
import sys
import pickle
import sys
from openpyxl import Workbook


savingPath = r"C:\python_projects\performanceTracker\times.pickle"
savingPathExcel = r"C:\python_projects\performanceTracker\time_table.xlsx"
w = win32gui
windows = {}
switch = 0
# window titles we want to have in one line:
title1 = "Brave"
title2 = "Google Chrome"


def main():
    print("Program is running...(Type 'quit' to exit)")
    windowTime()


def turning_off():
    global switch

    while True:
        response = input("> ").lower().strip()

        if response == "quit":
            switch = 1
            time.sleep(2)

            # merge together desired output so we don't have too many lines in excel
            finalDict = mergeNames(windows, title1, title2)

            with open(savingPath, "wb") as f:
                pickle.dump(finalDict, f)
            sys.exit()


def mergeNames(titlesDict, title1, title2):
    sumUp = 0
    sumUp2 = 0

    # window titles usually end with the name of program e.g. [something] - Google Chrome
    for k in list(titlesDict):
        if k.endswith(title1):
            sumUp += titlesDict[k]
            del titlesDict[k]
        elif k.endswith(title2):
            sumUp2 += titlesDict[k]
            del titlesDict[k]

    if sumUp > 0:
        titlesDict[title1] = sumUp
    if sumUp2 > 0:
        titlesDict[title2] = sumUp2

    # sort dictionary in descending order
    sorted_dict = dict(sorted(titlesDict.items(), key=lambda item: item[1], reverse=True))

    # convert seconds in dictionary to minutes
    for i in sorted_dict:
        sorted_dict[i] = str(datetime.timedelta(seconds=sorted_dict[i]))

    return sorted_dict


def windowTime():
    # get active window
    currentWindow = w.GetWindowText(w.GetForegroundWindow())

    while True:
        # check for window
        checkWindow = w.GetWindowText(w.GetForegroundWindow())

        # if there is a change of windows, change currentWindow and make a dictionary
        if checkWindow != currentWindow:
            currentWindow = w.GetWindowText(w.GetForegroundWindow())
            windows.setdefault(currentWindow, 0)

            # if there is no change, count seconds and write it to "windows" dictionary
            # always watching switch, serves for termination
            while w.GetWindowText(w.GetForegroundWindow()) == currentWindow:
                time.sleep(1)
                windows[currentWindow] += 1
                if switch == 1:
                    return


def write_to_excel():
    with open(savingPath, "rb") as f:
        timeDict = pickle.load(f)

    # make an Excel file
    wb = Workbook()
    ws = wb.active

    # make list of dict. keys
    k = list(timeDict)

    for i in range(len(timeDict)):
        ws.column_dimensions["A"].width = 70
        # first cell will be date and time of saving
        t = datetime.datetime.now()
        currentDateTime = t.strftime("%d.%m.%y - %H:%M:%S")
        ws["A1"] = currentDateTime
        # writing data to excel file
        ws.cell(row=i + 2, column=1).value = k[i]
        ws.cell(row=i + 2, column=2).value = timeDict[k[i]]
    wb.save(savingPathExcel)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "save":
        write_to_excel()
    else:
        threading.Thread(target=turning_off).start()
        main()
